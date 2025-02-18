import requests
from openpyxl import Workbook ,load_workbook
import pandas as pd

API_TOKEN = "7ef7c059-8118-4753-9031-e9f84b914447"  # This API key is only for testing purposes
BASE_URL = "https://api.sandbox.transferwise.tech"  # Use: https://api.wise.com for real data
header = {"Authorization": f"Bearer {API_TOKEN}"}

def connect(endpoint:str) -> list:
    response = requests.get(f"{BASE_URL}{endpoint}", headers=header)
    print("Connected") if response.status_code == 200 else print(f"Error Connecting: {response.status_code}")
    return response.json()

def flatten(data:list|dict):
    filtered=[]
    def walk(data, filtered):
        for d in data.items():
            items.append(d) if not isinstance(d[1], dict or list) else walk(d[1], filtered)
    for d in data:
         items=[]
         walk(d, items)
         filtered.append(dict(items))
    return filtered

def process(data:list, keys:tuple|list) -> list:
    processed:list = []
    for d in data:
        new_entry:dict = {}
        for key in keys:
            new_entry[key] = d[key]
        processed.append(new_entry)
    return processed


# Connecting and retrieving data
profile = connect("/v1/profiles")
p_id = profile[0]["id"]

activity = connect(f"/v1/profiles/{p_id}/activities")
balances = connect(f"/v4/profiles/{p_id}/balances?types=STANDARD")



# Formatting and pre-processing data
keepA = ("type", "description","primaryAmount", "secondaryAmount", "status", "createdOn", "updatedOn")
del activity['cursor']
activity = activity['activities']
activity = flatten(activity)
activity = process(activity, keepA)

keepB = ("currency", "value", "modificationTime")
balances = flatten(balances)
balances = process(balances, keepB)


# Writing data to excel
try: 
    file = load_workbook("wise.xlsx")
except FileNotFoundError:
    wb = Workbook()
    wb['Sheet'].title = "Activity"
    for col, val in enumerate(keepA):
        wb['Activity'].cell(row=1, column=col+1, value=val)
    wb.create_sheet("Balances")
    for col, val in enumerate(keepB):
        wb['Balances'].cell(row=1, column=col+1, value=val)
    wb.save("wise.xlsx")
    file = load_workbook("wise.xlsx")

fileA = file['Activity']
fileB = file['Balances']

idxA = fileA.max_row
idxB= fileB.max_row

with pd.ExcelWriter("wise.xlsx", mode='a', if_sheet_exists='overlay') as writer:
    pd.DataFrame(activity).to_excel(writer, sheet_name="Activity", startrow=idxA, header=False, index=False)
    pd.DataFrame(balances).to_excel(writer, sheet_name="Balances", startrow=idxB, header=False, index=False)
    


print("SUCCESS")